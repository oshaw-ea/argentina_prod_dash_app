import pandas as pd
from typing import List, Any

from helper_functions_ea import EnergyProduct
from openpyxl.styles import PatternFill
from openpyxl.worksheet.protection import SheetProtection
from argentina_prod.configs.enums import ModelMetadata
from typing import Dict, Optional
from io import BytesIO
import io, copy, base64
from dash import dcc
from datetime import datetime
from argentina_prod.configs.enums import ShoojuFields
from argentina_prod.pipeline.global_pipeline import GlobalPipeline


class PagesNames:
    MODELING = 'Modeling'


class PagesOrder:
    MODELING = 0


class PagesPrefix:
    MODELING = 'mod-'


class DataCategories:
    RIG_COUNT = 'rig_count'
    DRILLED_WELLS = 'drilled_wells'
    COMPLETIONS = 'completions'
    PERMITS = 'permits'
    DUCS = 'ducs'
    PRODUCTION = 'production'


class ChartTitles:
    DRILLED_WELLS = "Drilled Wells"
    COMPLETIONS = "Completions"
    PERMITS = "Permits"
    DUCS = "DUCs"
    PRODUCTION = "Production (mb/d)"

class Enums:
    IS_FORECAST = "is_forecast"
    METADATA = "metadata"


def pack_data(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df is None:
        return []
    if isinstance(df.index, pd.DatetimeIndex):
        df = df.copy()
        df.index = df.index.strftime('%Y-%m')
    return df.reset_index().to_dict('records')


def unpack_data(data: list) -> pd.DataFrame:
    if data is None:
        df = None
    elif len(data) == 0:
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(data).set_index(ShoojuFields.DT)
        df.index = pd.to_datetime(df.index)
        df.index = df.index.tz_localize(None)
        for col in df.columns:
            try:
                if col == ModelMetadata.VINTAGE:
                    df[col] = df[col].astype(int)
                else:
                    df[col] = df[col].astype(float)
            except (ValueError, TypeError):
                pass
    return df

def unpack_total_data(total_data: list) -> pd.DataFrame:
    if total_data is None:
        df = None
    elif len(total_data) == 0:
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(total_data)

    return df

def is_initial_load(table_data: pd.DataFrame, previous_table_data: pd.DataFrame):
    """When we load data in the tables we do not want to update them
    If previous data is None, it means it's the first time the page loads
    If columns don't match it means we swapped basin so it's an initial load too
    If columns match we are iterating over the table -> the user did update the table
    """
    if previous_table_data is None:
        return True
    if set(table_data.columns) != set(previous_table_data.columns):
        return True
    return False

def get_modified_plays_wide_df(previous_df, new_df) -> List:
    modified_columns = []
    for col in new_df.columns:
        if not previous_df[col].equals(new_df[col]):
            modified_columns.append(col)
    return modified_columns


# def get_modified_plays_long_df(previous_df, new_df) -> List:
#     modified_plays = []
#
#     for play in previous_df[ModelMetadata.EA_PLAY].unique():
#         prev_vals = previous_df[previous_df[ModelMetadata.EA_PLAY] == play][ModelMetadata.VAL].round(0)
#         new_vals = new_df[new_df[ModelMetadata.EA_PLAY] == play][ModelMetadata.VAL].round(0)
#
#         if not prev_vals.equals(new_vals):
#             modified_plays.append(play)
#
#     return modified_plays


def update_forecast_with_new_data(
        old_forecast: pd.DataFrame,
        new_data: pd.DataFrame,
        ) -> pd.DataFrame:
    new_forecast = old_forecast.copy()

    for play in new_data.columns:
        valid_data = new_data[play].dropna()
        play_date_mask = (new_forecast[ModelMetadata.EA_PLAY] == play) & (new_forecast.index.isin(valid_data.index))
        new_forecast.loc[play_date_mask, ModelMetadata.VAL] = valid_data

    # modified_plays = get_modified_plays_long_df(
    #     previous_df=old_forecast,
    #     new_df=new_forecast
    # )
    # return new_forecast[new_forecast[ModelMetadata.EA_PLAY].isin(modified_plays)]

    return new_forecast


def download_excel(
        dataframes: Dict[str, pd.DataFrame],
        filename_prefix: str,
        identifiers: Optional[Dict[str, str]] = None,
        timestamp: bool = True
) -> Dict:
    if not dataframes:
        return None

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            df = sort_dataframe_columns_in_specific_order(df)
            df.to_excel(writer, sheet_name=sheet_name, index=True)

            if hasattr(df, Enums.IS_FORECAST) and isinstance(df.is_forecast, pd.Series):
                worksheet = writer.sheets[sheet_name]
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                forecast_rows = df[df.is_forecast].index

                for row_idx, _ in enumerate(df.index, start=2):
                    if df.index[row_idx - 2] in forecast_rows:
                        for col_idx in range(1, len(df.columns) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill
        if identifiers:
            add_metadata_sheet(writer, identifiers)

    filename_parts = [filename_prefix]
    if identifiers:
        filename_parts.extend(f"{v}" for v in identifiers.values())

    if timestamp:
        filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))

    filename = '_'.join(filename_parts) + '.xlsx'

    return dcc.send_bytes(output.getvalue(), filename)


def sort_dataframe_columns_in_specific_order(df: pd.DataFrame) -> pd.DataFrame:
    def is_conventional_column(col):
        return 'conventional' in col.lower()

    conventional_columns = sorted([col for col in df.columns if is_conventional_column(col)])
    special_columns = []

    if 'Total' in df.columns:
        special_columns.append('Total')
    if Enums.IS_FORECAST in df.columns:
        special_columns.append(Enums.IS_FORECAST)

    other_columns = sorted([
        col for col in df.columns
        if not is_conventional_column(col)
    ])

    ordered_columns = other_columns + conventional_columns + special_columns

    return df[ordered_columns]

def add_metadata_sheet(writer, identifiers):
    metadata_df = pd.DataFrame({
        'Parameter': identifiers.keys(),
        'Value': identifiers.values()
    })

    metadata_df.to_excel(writer, sheet_name=Enums.METADATA, index=False)

    workbook = writer.book
    metadata_sheet = workbook[Enums.METADATA]

    metadata_sheet.sheet_state = 'hidden'
    metadata_sheet.protection = SheetProtection(
        sheet=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        sort=False,
        autoFilter=False,
        pivotTables=False
    )

# def extract_excel_upload(contents, filename, current_data, table_name, expected_basin):
#     content_type, content_string = contents.split(',')
#     decoded = base64.b64decode(content_string)
#
#     error_message = ""
#     error_status = False
#
#     eia_adjustment_only = False
#     if table_name == DataCategories.PRODUCTION:
#         eia_adjustment_only = True
#
#     try:
#         if 'xls' not in filename:
#             raise ValueError(f'{filename} not an excel file')
#         excel_file = pd.ExcelFile(io.BytesIO(decoded))
#         if Enums.METADATA not in excel_file.sheet_names:
#             raise ValueError(f'{Enums.METADATA} sheet not in {filename}.')
#         metadata = pd.read_excel(excel_file, sheet_name=Enums.METADATA)
#         excel_upload_basin = metadata[metadata["Parameter"]==ModelMetadata.EIA_BASIN].iloc[0]["Value"]
#         if excel_upload_basin != expected_basin:
#             raise ValueError(f'Excel data is for {excel_upload_basin.title()}, expected {expected_basin.title()}.')
#
#         if table_name not in excel_file.sheet_names:
#             raise ValueError(f'{table_name} sheet not in {filename}')
#         df = pd.read_excel(io.BytesIO(decoded), sheet_name=table_name)
#
#         forecast_rows = df[df[Enums.IS_FORECAST] == True]
#         if forecast_rows.empty:
#             raise ValueError(f'No {Enums.IS_FORECAST} rows in {filename}.')
#
#         excel_forecast_index = forecast_rows[ModelMetadata.DT].to_list()
#         table_forecast_index = [pd.Timestamp(row[ShoojuFields.DT]) for row in current_data]
#
#         if len(excel_forecast_index) > 0:
#             excel_start_date = excel_forecast_index[0]
#
#             relevant_table_dates = [date for date in table_forecast_index if date >= excel_start_date]
#
#             for i, table_date in enumerate(relevant_table_dates):
#                 if i >= len(excel_forecast_index) or excel_forecast_index[i] != table_date:
#                     raise ValueError(f'Expected date {table_date} missing from forecast rows in {filename}.')
#
#         updated_data = copy.deepcopy(current_data)
#         date_to_index = {pd.Timestamp(row[ShoojuFields.DT]): i for i, row in enumerate(updated_data)}
#         for _, forecast_row in forecast_rows.iterrows():
#             date = forecast_row[ShoojuFields.DT]
#             if date in date_to_index:
#                 idx = date_to_index[date]
#                 for col in forecast_row.index:
#                     if eia_adjustment_only:
#                         if ModelMetadata.EIA_ADJUSTMENT_SERIES in col and col in updated_data[idx]:
#                             updated_data[idx][col] = forecast_row[col]
#                     else:
#                         if col != ShoojuFields.DT and col in updated_data[idx]:
#                             updated_data[idx][col] = forecast_row[col]
#
#         if updated_data == current_data:
#             raise ValueError(f'{table_name} data in {filename} is the same as that already in the model.')
#
#         return updated_data, current_data, error_message, error_status
#
#     except Exception as e:
#         error_message = str(e)
#         error_status = True
#         return current_data, False, error_message, error_status


def make_continuous_df(history_df, forecast_df) -> pd.DataFrame:
    df = pd.concat([history_df, forecast_df], axis=0)
    df = df[~df.index.duplicated(keep='first')]

    return df


def make_continuous_df_with_forecast_flag(history_df, forecast_df) -> pd.DataFrame:
    df = pd.concat([history_df, forecast_df], axis=0)
    df = df[~df.index.duplicated(keep='first')]

    df[Enums.IS_FORECAST] = False
    if forecast_df is not None and not forecast_df.empty:
        df.loc[forecast_df.index, Enums.IS_FORECAST] = True

    return df

# def melt_and_add_fields(df: pd.DataFrame, basin: str, energy_product: str = EnergyProduct().CRUDE_OIL) -> pd.DataFrame:
#     df = df.melt(var_name=ModelMetadata.EA_PLAY, value_name=ShoojuFields.VAL,
#                          ignore_index=False)
#     df[ModelMetadata.EIA_BASIN] = basin
#     df[ModelMetadata.ENERGY_PRODUCT] = energy_product
#     df = df.dropna()
#
#     return df

# def melt_and_add_fields_rigs_baseline(df: pd.DataFrame) -> pd.DataFrame:
#     df = df.melt(var_name=ModelMetadata.EA_PLAY, value_name=ShoojuFields.VAL,
#                  ignore_index=False)
#     locations = UsBasins().mapping[[ModelMetadata.EA_PLAY, ModelMetadata.EIA_BASIN]]
#     df = df.reset_index().merge(locations, on=ModelMetadata.EA_PLAY, how='left').set_index(ModelMetadata.DT)
#     df[ModelMetadata.EIA_BASIN] = df.apply(
#         lambda row: row[ModelMetadata.EA_PLAY]
#         if row[ModelMetadata.EIA_BASIN] == UsBasins.SINGLEPLAY_METADATA_INDICATOR
#         else row[ModelMetadata.EIA_BASIN],
#         axis=1
#     )
#
#     df[ModelMetadata.ENERGY_PRODUCT] = EnergyProduct().CRUDE_OIL
#
#     return df
#
# def melt_and_add_fields_rig_price_sensitive(df: pd.DataFrame, basin: str) -> pd.DataFrame:
#     df = df.melt(var_name=ModelMetadata.EA_PLAY, value_name=ShoojuFields.VAL,
#                  ignore_index=False)
#     df[ModelMetadata.EIA_BASIN] = basin
#     df[RigCountMetadata.RIG_COUNT_TYPE] = RigCountMetadata.PRICE_SENSITIVE
#     df[ModelMetadata.PRICE_SCENARIO] = PriceScenario.EA_FORECAST
#     df = df.dropna()
#     df[ModelMetadata.ENERGY_PRODUCT] = EnergyProduct().CRUDE_OIL
#
#     return df

def prioritise_df(priority_df: pd.DataFrame, non_priority_df: pd.DataFrame, match_columns: list = None) -> pd.DataFrame:
    priority_copy = priority_df.copy()
    non_priority_copy = non_priority_df.copy()

    if match_columns:
        priority_copy = priority_copy.set_index(match_columns, append=True)
        non_priority_copy = non_priority_copy.set_index(match_columns, append=True)

        combined_df = priority_copy.combine_first(non_priority_copy)

        combined_df = combined_df.reset_index(level=match_columns)
    else:
        combined_df = priority_copy.combine_first(non_priority_copy)

    return combined_df


# def add_padd(df: pd.DataFrame):
#     if ModelMetadata.PADD in df.columns:
#         return df
#     locations = UsBasins().mapping
#     locations = locations[[ModelMetadata.PADD, ModelMetadata.EA_PLAY]]
#
#
#     result = pd.merge(
#         df.reset_index(),
#         locations,
#         on=ModelMetadata.EA_PLAY,
#         how='left'
#     ).set_index(ModelMetadata.DT)
#
#     return result




def preserve_figure_state(new_fig, current_figure):
    if current_figure is not None:
        current_layout = current_figure['layout']

        for axis in ['xaxis', 'yaxis']:
            if axis in current_layout:
                new_fig.update_layout(**{
                    axis: dict(
                        range=current_layout[axis].get('range'),
                        autorange=current_layout[axis].get('autorange'),
                        type=current_layout[axis].get('type'),
                        scaleanchor=current_layout[axis].get('scaleanchor'),
                        scaleratio=current_layout[axis].get('scaleratio')
                    )
                })

        if 'legend' in current_layout:
            new_fig.update_layout(
                showlegend=current_layout['legend'].get('showlegend'),
                legend=dict(
                    x=current_layout['legend'].get('x'),
                    y=current_layout['legend'].get('y'),
                    orientation=current_layout['legend'].get('orientation')
                )
            )

        if 'data' in current_figure:
            total_trace_old = next((trace for trace in current_figure['data'] if trace['name'] == 'Total'), None)
            total_trace_new = next((trace for trace in new_fig.data if trace.name == 'Total'), None)

            if total_trace_old and total_trace_new:
                for prop in ['visible', 'showlegend']:
                    if prop in total_trace_old:
                        setattr(total_trace_new, prop, total_trace_old[prop])

            for new_trace, old_trace in zip(new_fig.data, current_figure['data']):
                if 'name' in old_trace.keys():
                    if old_trace['name'] != 'Total':
                        for prop in ['visible', 'showlegend', 'legendgroup']:
                            if prop in old_trace:
                                setattr(new_trace, prop, old_trace[prop])

    return new_fig

def get_other_energy_product(energy_product):
    if energy_product == EnergyProduct().NATURAL_GAS:
        return EnergyProduct().CRUDE_OIL
    if energy_product == EnergyProduct().CRUDE_OIL:
        return EnergyProduct().NATURAL_GAS