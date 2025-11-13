import pandas as pd
from typing import List, Any

from helper_functions_ea import EnergyProduct
from openpyxl.styles import PatternFill
from openpyxl.worksheet.protection import SheetProtection
from argentina_prod.configs.enums import ModelMetadata
from typing import Dict, Optional
from io import BytesIO
import io, copy, base64
from dash import dcc
from datetime import datetime
from argentina_prod.configs.enums import ShoojuFields
from argentina_prod.pipeline.global_pipeline import GlobalPipeline


class PagesNames:
    MODELING = 'Modeling'


class PagesOrder:
    MODELING = 0


class PagesPrefix:
    MODELING = 'mod-'


class DataCategories:
    DRILLED_WELLS = 'drilled_wells'
    COMPLETIONS = 'completions'
    PRODUCTION = 'production'


class ChartTitles:
    DRILLED_WELLS = "Drilled Wells"
    COMPLETIONS = "Completions"
    PRODUCTION = "Production"

class Enums:
    IS_FORECAST = "is_forecast"
    METADATA = "metadata"


def pack_data(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df is None:
        return []
    if isinstance(df.index, pd.DatetimeIndex):
        df = df.copy()
        df.index = df.index.strftime('%Y-%m')
    return df.reset_index().to_dict('records')


def unpack_data(data: list) -> pd.DataFrame:
    if data is None:
        df = None
    elif len(data) == 0:
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(data).set_index(ShoojuFields.DT)
        df.index = pd.to_datetime(df.index)
        df.index = df.index.tz_localize(None)
        for col in df.columns:
            try:
                if col == ModelMetadata.VINTAGE:
                    df[col] = df[col].astype(int)
                else:
                    df[col] = df[col].astype(float)
            except (ValueError, TypeError):
                pass
    return df

def unpack_total_data(total_data: list) -> pd.DataFrame:
    if total_data is None:
        df = None
    elif len(total_data) == 0:
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(total_data)

    return df

def is_initial_load(table_data: pd.DataFrame, previous_table_data: pd.DataFrame):
    """When we load data in the tables we do not want to update them
    If previous data is None, it means it's the first time the page loads
    If columns don't match it means we swapped basin so it's an initial load too
    If columns match we are iterating over the table -> the user did update the table
    """
    if previous_table_data is None:
        return True
    if set(table_data.columns) != set(previous_table_data.columns):
        return True
    return False



def download_excel(
        dataframes: Dict[str, pd.DataFrame],
        filename_prefix: str,
        identifiers: Optional[Dict[str, str]] = None,
        timestamp: bool = True
) -> Dict:
    if not dataframes:
        return None

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            df = sort_dataframe_columns_in_specific_order(df)
            df.to_excel(writer, sheet_name=sheet_name, index=True)

            if hasattr(df, Enums.IS_FORECAST) and isinstance(df.is_forecast, pd.Series):
                worksheet = writer.sheets[sheet_name]
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                forecast_rows = df[df.is_forecast].index

                for row_idx, _ in enumerate(df.index, start=2):
                    if df.index[row_idx - 2] in forecast_rows:
                        for col_idx in range(1, len(df.columns) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill
        if identifiers:
            add_metadata_sheet(writer, identifiers)

    filename_parts = [filename_prefix]
    if identifiers:
        filename_parts.extend(f"{v}" for v in identifiers.values())

    if timestamp:
        filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))

    filename = '_'.join(filename_parts) + '.xlsx'

    return dcc.send_bytes(output.getvalue(), filename)


def sort_dataframe_columns_in_specific_order(df: pd.DataFrame) -> pd.DataFrame:
    def is_conventional_column(col):
        return 'conventional' in col.lower()

    conventional_columns = sorted([col for col in df.columns if is_conventional_column(col)])
    special_columns = []

    if 'Total' in df.columns:
        special_columns.append('Total')
    if Enums.IS_FORECAST in df.columns:
        special_columns.append(Enums.IS_FORECAST)

    other_columns = sorted([
        col for col in df.columns
        if not is_conventional_column(col)
    ])

    ordered_columns = other_columns + conventional_columns + special_columns

    return df[ordered_columns]

def add_metadata_sheet(writer, identifiers):
    metadata_df = pd.DataFrame({
        'Parameter': identifiers.keys(),
        'Value': identifiers.values()
    })

    metadata_df.to_excel(writer, sheet_name=Enums.METADATA, index=False)

    workbook = writer.book
    metadata_sheet = workbook[Enums.METADATA]

    metadata_sheet.sheet_state = 'hidden'
    metadata_sheet.protection = SheetProtection(
        sheet=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        sort=False,
        autoFilter=False,
        pivotTables=False
    )


def make_continuous_df(history_df, forecast_df) -> pd.DataFrame:
    df = pd.concat([history_df, forecast_df], axis=0)
    df = df[~df.index.duplicated(keep='first')]

    return df


def make_continuous_df_with_forecast_flag(history_df, forecast_df) -> pd.DataFrame:
    df = pd.concat([history_df, forecast_df], axis=0)
    df = df[~df.index.duplicated(keep='first')]

    df[Enums.IS_FORECAST] = False
    if forecast_df is not None and not forecast_df.empty:
        df.loc[forecast_df.index, Enums.IS_FORECAST] = True

    return df

def prioritise_df(priority_df: pd.DataFrame, non_priority_df: pd.DataFrame, match_columns: list = None) -> pd.DataFrame:
    priority_copy = priority_df.copy()
    non_priority_copy = non_priority_df.copy()

    if match_columns:
        priority_copy = priority_copy.set_index(match_columns, append=True)
        non_priority_copy = non_priority_copy.set_index(match_columns, append=True)

        combined_df = priority_copy.combine_first(non_priority_copy)

        combined_df = combined_df.reset_index(level=match_columns)
    else:
        combined_df = priority_copy.combine_first(non_priority_copy)

    return combined_df


def preserve_figure_state(new_fig, current_figure):
    if current_figure is not None:
        current_layout = current_figure['layout']

        for axis in ['xaxis', 'yaxis']:
            if axis in current_layout:
                new_fig.update_layout(**{
                    axis: dict(
                        range=current_layout[axis].get('range'),
                        autorange=current_layout[axis].get('autorange'),
                        type=current_layout[axis].get('type'),
                        scaleanchor=current_layout[axis].get('scaleanchor'),
                        scaleratio=current_layout[axis].get('scaleratio')
                    )
                })

        if 'legend' in current_layout:
            new_fig.update_layout(
                showlegend=current_layout['legend'].get('showlegend'),
                legend=dict(
                    x=current_layout['legend'].get('x'),
                    y=current_layout['legend'].get('y'),
                    orientation=current_layout['legend'].get('orientation')
                )
            )

        if 'data' in current_figure:
            total_trace_old = next((trace for trace in current_figure['data'] if trace['name'] == 'Total'), None)
            total_trace_new = next((trace for trace in new_fig.data if trace.name == 'Total'), None)

            if total_trace_old and total_trace_new:
                for prop in ['visible', 'showlegend']:
                    if prop in total_trace_old:
                        setattr(total_trace_new, prop, total_trace_old[prop])

            for new_trace, old_trace in zip(new_fig.data, current_figure['data']):
                if 'name' in old_trace.keys():
                    if old_trace['name'] != 'Total':
                        for prop in ['visible', 'showlegend', 'legendgroup']:
                            if prop in old_trace:
                                setattr(new_trace, prop, old_trace[prop])

    return new_fig

def get_other_energy_product(energy_product):
    if energy_product == EnergyProduct().NATURAL_GAS:
        return EnergyProduct().CRUDE_OIL
    if energy_product == EnergyProduct().CRUDE_OIL:
        return EnergyProduct().NATURAL_GAS